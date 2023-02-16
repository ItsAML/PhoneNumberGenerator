import openpyxl

# Enter The First Three Digit that you wanna make numbers with. Note: Dont Enter Zero Before Digits.
threenum = 935

# The Following Variable Below Defines How many numbers should be stored in a single xlsx file(The Maximum Number is: 120000 but its not gonna work in mobile if its more than 9999)
how_many_numbers = 9999


print("started...")
# start range and end range

start_range = int(f'{threenum}0000000')
end_range = int(f'{threenum}9900000')
distance = end_range - start_range

# calculate number of files needed
num_files = int(distance / how_many_numbers) + 1

# loop through each file
for file_num in range(num_files):
    # create a new workbook
    wb = openpyxl.Workbook()

    # select the first sheet
    sheet = wb.active

    # input names and numbers
    start = start_range + file_num * how_many_numbers
    end = start + how_many_numbers
    names = [i for i in range(1, how_many_numbers)]
    numbers = [f"0{i}" for i in range(start, end)]

    # write values in cells
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Phone Number'
    for i in range(2, len(names) + 2):
        sheet.cell(row=i, column=1, value=names[i-2])
        sheet.cell(row=i, column=2, value=numbers[i-2])

    # save the workbook
    wb.save(f"names_and_numbers_{threenum}_{file_num}.xlsx")

    print(f"{threenum}PART{file_num}: DONE...")

print('all completed successfully')

# @AMLDevelopment on telegram
# https://github.com/ItsAML
